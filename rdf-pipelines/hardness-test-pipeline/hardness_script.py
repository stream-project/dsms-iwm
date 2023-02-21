import os
from string import Template

from data2rdf.cli.abox_conversion import run_abox_pipeline_for_folder
from data2rdf.annotation_pipeline import Annotation_Pipeline
from openpyxl import Workbook, load_workbook
from rdflib import Namespace, URIRef


BASE_DIR = os.path.abspath(os.path.dirname(__file__))
FOLDER = {
    # inputs
    'sheets': os.path.join(BASE_DIR, "input", "data"),
    'mapping': os.path.join(BASE_DIR, "input", "mappings"),
    'abox': os.path.join(BASE_DIR, "input", "method-graph"),
    'tbox': os.path.join(BASE_DIR, "input", "ontology"),
    # intermediate
    'preprocessed_sheets': os.path.join(BASE_DIR, "input", "data", ".tmp"),
    'direct_output': os.path.join(BASE_DIR, "output", ".tmp"),
    # RDF output
    'final_output': os.path.join(BASE_DIR, "output"),
}
STREAM_IRI = "https://stream-projekt.net"
DATASET_API = "https://stream.materials-data.space/datasets"
DEBUG = os.getenv("DEBUG", "").capitalize() == "True"


def main():
    adapt_Brinell_datasheet_to_pipeline()

    # update the abox template whenever there is a change in the graph
    # one can set the argument to False if there is no need to update
    needs_update:bool = True
    if needs_update:
        update_tbox_template()

    # apply pipeline 
    apply_pipeline()

    # remove preprocessed sheets
    if not DEBUG:
        os.system(f"rm -r {FOLDER['preprocessed_sheets']}/*.xlsx")


def adapt_Brinell_datasheet_to_pipeline():
    """Preprocessing the excel sheets
    | combine "2 files per dataset" into "1 file with 2 sheets per dataset"
    | convert time column into ISO time format
    (with a lot of hard-coding)
    """

    # match all meta-data files by naming pattern
    data_files = filter(
        # NOTE hard-coded knowledge: naming pattern
        lambda f_name: f_name.startswith("STREAM_") and f_name.endswith("_IWM_BrinellIndentation.xlsx"),
        os.listdir(FOLDER['sheets'])
    )

    for f_name in data_files:

        # NOTE hard-coded knowledge: sheet names
        meta_data = load_workbook(os.path.join(FOLDER['sheets'], f_name))['user_variables']

        # NOTE sheet names are matching with what is specified in "location_mapping.xlsx"
        wb = Workbook()
        ws1 = wb.create_sheet("meta")
        for row in meta_data.rows:
            for cell in row:
                ws1[cell.coordinate] = cell.value
        _ = wb.create_sheet("columns")
        # harness test has no column data to fill in

        # save the new sheets
        wb.save(os.path.join(FOLDER['preprocessed_sheets'], f_name))

        # debug with only one dataset
        if DEBUG:
            break



def update_tbox_template():
    [tbox_file] = filter(lambda f: f.endswith(".ttl"), os.listdir(FOLDER['tbox']))
    tbox_file = os.path.join(FOLDER['tbox'], tbox_file)
    run_abox_pipeline_for_folder(FOLDER['abox'], onto_path=tbox_file)


def apply_pipeline():
    # create a pipeline object
    parser = "excel"

    location_mapping = os.path.join(FOLDER['mapping'], "location_mapping.xlsx")
    parser_args = {
        "location_mapping_f_path": location_mapping
    }

    [abox_raw] = filter(lambda f: f.endswith(".xml"), os.listdir(FOLDER['abox']))
    abox_template = os.path.join(FOLDER['abox'], abox_raw.replace(".xml", ""), abox_raw.replace(".xml", ".mod.ttl"))

    [tbox_file] = filter(lambda f: f.endswith(".ttl"), os.listdir(FOLDER['tbox']))
    tbox_file = os.path.join(FOLDER['tbox'], tbox_file)

    label_mapping = os.path.join(FOLDER['mapping'], "label_mapping.xlsx")

    pipeline = Annotation_Pipeline(
        None,  # to be specified iteratively later
        parser,
        parser_args,
        abox_template,
        label_mapping,
        FOLDER['direct_output'],
        ontology_file = tbox_file,
        base_iri = STREAM_IRI
    )

    # for entry in os.scandir(FOLDER['preprocessed_sheets']):
    for entry in os.scandir(FOLDER['preprocessed_sheets']):
        # check if file is valid
        title, file_format = os.path.splitext(entry.name)
        if file_format != ".xlsx":
            raise Warning("source folder not pure.")

        # run pipeline
        pipeline.input_file = entry.path
        pipeline.run_pipeline()
        graph = pipeline.export_graph()
    
        # a local fix for the problem that some namespaces doesn't have a proper prefix
        graph.bind("skos", "http://www.w3.org/2004/02/skos/core#")
        graph.bind("emmo", "http://emmo.info/emmo#")
        graph.bind("matvoc", "http://stream-ontology.com/matvoc-core/")

        # post-process: add connection to metadata annotations
        dataset_id = title.replace("STREAM_", "").replace("_IWM_BrinellIndentation", "")
        graph.add((
            Namespace(pipeline.file_uri).HardnessTestMeasurement,
            URIRef("http://stream-ontology.com/matvoc-core/hasMetadata"),
            URIRef(f"{DATASET_API}/{dataset_id}")
        ))

        # generate RDF output
        graph.serialize(os.path.join(FOLDER['final_output'], title+".ttl"), format="turtle")

        # debug with only one dataset
        if DEBUG:
            break


if __name__ == "__main__":
    main()
