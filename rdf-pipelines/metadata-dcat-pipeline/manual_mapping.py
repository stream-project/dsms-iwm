from datetime import date, datetime
from typing import Dict

from openpyxl.worksheet.worksheet import Worksheet  # for type hints
from simphony_osp.ontology import OntologyIndividual  # for type hints
from simphony_osp.namespaces import owl, dcat2, dcterms, foaf, vcard
from pytz import timezone


WEB_URL = "https://dsms.stream-dataspace.net"


def map_hardness_meta(ws: Worksheet) -> Dict:
    default_person = create_default_person()

    default_contact = create_default_contact()

    pot = dcterms.PeriodOfTime()
    pot_annotation = {
        dcat2.startDate: ws["B5"].value if ws["B5"].value else date.today().isoformat(),
        dcat2.endDate: ws["B6"].value if ws["B6"].value else date.today().isoformat(),
    }
    bind_annotation(pot, pot_annotation)

    distrib = dcat2.Distribution()
    distrib_annotation = {
        dcterms.title: f"STREAM_{ws['B2'].value}_IWM_BrinellIndentation.xlsx",
        dcat2.downloadURL: owl.Thing(iri=f"{WEB_URL}/iwm-data-files/STREAM_{ws['B2'].value}_IWM_BrinellIndentation.xlsx"),
        dcat2.mediaType: dcterms.MediaType(iri="https://www.iana.org/assignments/media-types/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    }
    bind_annotation(distrib, distrib_annotation)

    time_now = datetime.today().astimezone(timezone("Europe/Berlin")).isoformat()
    dataset_annotation = {
        dcterms.identifier: ws["B2"].value,
        dcterms.title: f"STREAM_{ws['B2'].value}_IWM_BrinellIndentation.xlsx",
        dcterms.description: "This is an indentation test dataset.",
        dcterms.issued: time_now,
        dcterms.modified: ws["B6"].value if ws["B6"].value else time_now,
        dcterms.temporal: pot,
        dcterms.publisher: default_person,
        dcterms.creator: create_person_with_nick(ws["B4"].value) if ws["B4"].value else default_person,
        dcat2.contactPoint: default_contact,
        dcterms.language: dcterms.LinguisticSystem(iri="http://id.loc.gov/vocabulary/iso639-1/en"),
        dcterms.license: dcterms.RightsStatement(iri="https://creativecommons.org/licenses/by-nc-nd/4.0/"),
        dcat2.landingPage: foaf.Document(iri="http://dsms.stream-dataspace.net/"),
        dcat2.distribution: distrib,
    }

    return dataset_annotation


def map_tensile_meta(ws: Worksheet) -> Dict:
    default_person = create_default_person()

    default_contact = create_default_contact()

    pot = dcterms.PeriodOfTime()
    pot_annotation = {
        dcat2.startDate: ws["B8"].value if ws["B8"].value else date.today().isoformat(),
        dcat2.endDate: ws["B9"].value if ws["B9"].value else date.today().isoformat(),
    }
    bind_annotation(pot, pot_annotation)

    distrib = dcat2.Distribution()
    distrib_annotation = {
        dcterms.title: f"STREAM_IWM_TT_{ws['B2'].value}",
        dcat2.downloadURL: owl.Thing(iri=f"{WEB_URL}/iwm-data-files/STREAM_IWM_TT_{ws['B2'].value}.xlsx"),
        dcat2.mediaType: dcterms.MediaType(iri="https://www.iana.org/assignments/media-types/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    }
    bind_annotation(distrib, distrib_annotation)

    time_now = datetime.today().astimezone(timezone("Europe/Berlin")).isoformat()
    dataset_annotation = {
        dcterms.identifier: ws["B2"].value,
        dcterms.title: f"STREAM_{ws['B2'].value}_IWM_TensileTest",
        dcterms.description: "This is a tensile test dataset.",
        dcterms.issued: time_now,
        dcterms.modified: ws["B9"].value if ws["B9"].value else time_now,
        dcterms.temporal: pot,
        dcterms.publisher: default_person,
        dcterms.creator: default_person,
        dcat2.contactPoint: default_contact,
        dcterms.language: dcterms.LinguisticSystem(iri="http://id.loc.gov/vocabulary/iso639-1/en"),
        dcterms.license: dcterms.RightsStatement(iri="https://creativecommons.org/licenses/by-nc-nd/4.0/"),
        dcat2.landingPage: foaf.Document(iri="http://dsms.stream-dataspace.net/"),
        dcat2.distribution: distrib,
    }

    return dataset_annotation


# def map_tensile_raw(ws: Worksheet) -> Dict:
#     manual_mapping = {
#         "Zeit": oboe_characteristics.Time(),
#         "[s]": oboe_standards.Second(),
#         "Länge": oboe_characteristics.Length(),
#         "[mm]": oboe_standards.Millimeter(),
#         "Kraft": oboe_characteristics.Force(),
#         "[kN]": stream.Kilonewton(),
#         "Dehnrate": stream.LengthPerLengthRate(),
#         "[-/s]": stream.MillimeterPerMillimeterPerSecond(),
#         "Dehnung": stream.LengthPerLength(),
#         "[-]": stream.MillimeterPerMillimeter(),
#         "Spannung": oboe_characteristics.Pressure(),
#         "[MPa]": stream.Megapascal(),
#     }

#     observ_set = set()
#     START_ROW = 8
#     END_ROW = (START_ROW + 5) if DEBUG else (ws.max_row + 1)
#     # each row is an observation
#     for value_row in range(START_ROW, END_ROW):
#         observation = oboe_core.Observation()

#         # each physical quantity of its row is a measurement
#         for column in range(ord("B"), ord("G") + 1):
#             measurement = oboe_core.Measurement()
#             observation.connect(measurement, rel=oboe_core.hasMeasurement)

#             raw_value = ws[f"{chr(column)}{value_row}"].value
#             measured_value = oboe_core.MeasuredValue(hasCode=raw_value)
#             measurement[oboe_core.hasValue] = measured_value

#             NAME_ROW = 5
#             raw_name = ws[f"{chr(column)}{NAME_ROW}"].value
#             characteristic = manual_mapping[raw_name]
#             measurement[oboe_core.ofCharacteristic] = characteristic

#             UNIT_ROW = 7
#             raw_unit = ws[f"{chr(column)}{UNIT_ROW}"].value
#             standard = manual_mapping[raw_unit]
#             measurement[oboe_core.usesStandard] = standard

#         observ_set |= {observation}
    
#     return {oboe_core.hasMember: observ_set}


''' following are some util functions
'''

def create_default_person() -> foaf.Person:
    default_person = foaf.Person()
    person_annotation = {
        foaf.familyName: "Weber",
        foaf.givenName: "Matthias",
        foaf.mbox: dcterms.Agent(iri="mailto:matthias.weber@iwm.fraunhofer.de"),
        foaf.nick: "wbr",
    }
    bind_annotation(default_person, person_annotation)
    return default_person


def create_default_contact() -> vcard.Contact:
    default_contact = vcard.Contact()
    contact_annotation = {
        vcard.__getattr__("family-name"): "Weber",
        vcard.__getattr__("given-name"): "Matthias",
        vcard.hasEmail: foaf.Agent(iri="mailto:matthias.weber@iwm.fraunhofer.de"),
        vcard.nickname: "wbr",
        # here organization is not created as a variable to be handed to wrapper
        # otherwise it will complain that the same iri has been assigned
        vcard.hasOrganizationName: vcard.Organization(
            iri="https://www.iwm.fraunhofer.de"
        ),
    }
    bind_annotation(default_contact, contact_annotation)
    return default_contact


def create_person_with_nick(nick: str) -> foaf.Person:
    person = foaf.Person()
    person_annotation = {
        foaf.nick: nick,
    }
    bind_annotation(person, person_annotation)
    return person


def bind_annotation(subj: OntologyIndividual, description: Dict) -> None:
    '''a function to annotate the subject, regardless of whether the predicate 
    is object property or datatype property, functional or non-functional.
    String literals is automatically converted to individuals of the type cuba.Entity,
    because osp-core accepts only OntologyIndividual object.

    subj -- subject
    description -- Dict[predicate, object]
    '''

    for pred, obj in description.items():
        subj[pred] += obj
